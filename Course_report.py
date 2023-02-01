from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl import workbook, load_workbook
import time

wb = openpyxl.load_workbook("Course Report Scrape.xlsx")
sheet1 = wb['Edited']
wb.create_sheet("Contact")
sh = wb["Contact"]
sh['A1'].value = "Company Name"
sh['B1'].value = "Website"
sh['C1'].value = "Mail"
sh['D1'].value = "Social Media"

row = sheet1.max_row
column = sheet1.max_column

for i in range(2,row+1):
    for j in range(1, column + 1):
        company = sheet1.cell(i, j).value
        #print(company)
        html_link = requests.get(f'{company}').text
        soup = BeautifulSoup(html_link, 'html.parser')
        #Name, URL, Mail id, course, location, linkedin profile
        ul_class = soup.find("ul" , class_ = "school-info")
        #To find the tracks
        #print(ul_class)
        li_tracks_class = ul_class.find("li", class_ = "school-tracks text-left")
        #print(li_tracks_class)
        a_tags = li_tracks_class.find_all("a")
        #print(a_tags)
        a_tags_lst=[]
        for tag in a_tags:
            a_tags_lst.append(tag.text)
        print(a_tags_lst)
        # To find company url
        ul_li_url_name = ul_class.find("li", class_ ="url text-center-desktop-only")
        url_tag = ul_li_url_name.a
        sh.cell(i,j+1).value = url_tag.get("href").split("?")[0].rstrip("/")
        print(url_tag.get("href").split("?")[0].rstrip("/"))
        # To find company Name
        sh.cell(i, j).value = url_tag.text
        #To find company mail
        ul_li_mail = ul_class.find("li", class_ = "email text-center-desktop-only")
        if ul_class.find("li", class_ = "email text-center-desktop-only") == None:
            pass
        else:
            if ul_li_mail.a == None:
                pass
            else:
                mail_tag = ul_li_mail.a
                sh.cell(i, j+2).value = mail_tag.get("href").split(":")[1]
                print(mail_tag.get("href").split(":")[1])
        #To find linkedin profile
        if len(soup.find_all("a", class_="no-decoration")) < 1:
            pass
        else:
            socialmedia_tag = soup.find_all("a", class_="no-decoration")
            for tag in socialmedia_tag:
                #print(tag)
                if "linkedin" in tag.get("href"):
                    sh.cell(i,j+3).value = tag.get("href")
                    linkedin_link = tag.get("href")
                    print(linkedin_link)
                    print()
wb.save("Course Report Scrape.xlsx")










